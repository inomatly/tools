import openpyxl
import datetime

def get_column_index_for_day(day):
    days = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
    return 7 + days.index(day), 15 + days.index(day), 23 + days.index(day)

def extract_tasks(sheet, day_col):
    tasks = []
    for row in range(3, 27):
        category = sheet.cell(row=row, column=1).value
        task = sheet.cell(row=row, column=2).value
        planned_time = sheet.cell(row=row, column=day_col[0]).value
        actual_time = sheet.cell(row=row, column=day_col[1]).value
        if (not planned_time or planned_time == 0) and (not actual_time or actual_time == 0):
            continue
        tasks.append((category, task, planned_time, actual_time))
    return tasks

def extract_objectives(sheet, day_col):
    objectives = {}
    for row in range(3, 27):
        task = sheet.cell(row=row, column=2).value
        objective = sheet.cell(row=row, column=day_col[2]).value
        if objective:
            objectives[task] = objective
    return objectives

def extract_completed(sheet, day_col):
    completed = [sheet.cell(row=row, column=day_col[2]).value for row in range(33, 37) if sheet.cell(row=row, column=day_col[2]).value]
    return completed

def extract_incomplete(sheet, day_col):
    incomplete = [sheet.cell(row=row, column=day_col[2]).value for row in range(37, 41) if sheet.cell(row=row, column=day_col[2]).value]
    return incomplete

def generate_report(filename):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.active

    today = datetime.datetime.today().strftime('%A')[:3]
    day_col = get_column_index_for_day(today)

    tasks = extract_tasks(sheet, day_col)
    objectives = extract_objectives(sheet, day_col)
    completed = extract_completed(sheet, day_col)
    incomplete = extract_incomplete(sheet, day_col)

    with open(f"日報{datetime.datetime.now().strftime('%m%d')}.txt", 'w', encoding="utf-8") as f:
        f.write("\t\t\t\t成果物目標\n")
        for task in tasks:
            f.write(f"{task[0]}:{task[1]}\t\t{task[3] if task[3] else 0}min\t({task[2] if task[2] else 0}min)\n")
            if task[1] in objectives:
                f.write(f"{objectives[task[1]]}\n")
        f.write("\n完了\n")
        for item in completed:
            f.write(f"・ {item}\n")
        f.write("\n未完了\n")
        for item in incomplete:
            f.write(f"・ {item}\n")

if __name__ == '__main__':
    import sys
    if len(sys.argv) != 2:
        print("Usage: python transrate.py <input_file>")
        sys.exit(1)
    generate_report(sys.argv[1])
